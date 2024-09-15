import math

import numpy as np
import open3d as o3d
import heapq


def detect_outliers2(df):
    outlier_indices = []
    # 1st quartile (25%)
    Q1 = np.percentile(df, 25)
    # print( Q1)
    # 3rd quartile (75%)
    Q3 = np.percentile(df, 75)
    # Interquartile range (IQR)
    IQR = Q3 - Q1

    # outlier step
    outlier_step = 1.5 * IQR
    # print(outlier_step)
    # print(len(df))
    i=0
    d=[]
    for nu in df:
        i=i+1
        if (nu > Q1 - outlier_step) and (nu < Q3 + outlier_step):
            d.append(nu)
        # print(i)

    return d

def leaf_high_width(pcd):
    '''

    :param pcd:Leaf point cloud after PCA
    :return: The leaves are long and wide
    '''

    obb = pcd.get_oriented_bounding_box()

    # 获取OBB的长和宽
    extent = obb.extent
    length = extent[0]
    width = extent[1]
    return length,width

import os
from sklearn.decomposition import PCA
name=['ID']
lenth = []
width = []
data_r='D:\datains\论文data\分割结果\End2End白盆'
for itemy in os.listdir(data_r):
    # print(itemy.split('.')[0])
    name.append(itemy)
    num_b=1
    xyzl=np.loadtxt(os.path.join(data_r,itemy))
    data=xyzl[xyzl[:,-2]==2]
    max=np.max(data[:,-1])
    min=np.min(data[:,-1])
    l=np.unique(data[:,-1])
    h1=[]
    w1=[]
    idx=[]
    for i in l:
        num_b+=1
        pcd=o3d.geometry.PointCloud()
        pcd.points=o3d.utility.Vector3dVector(data[data[:,-1]==float(i)][:,:3])

        if len(np.array(pcd.points)) < 10:

            continue

        X = np.asarray(pcd.points)
        pca = PCA(n_components=3)
        # print(X)
        pca.fit(X)
        X_pca = pca.transform(X)


        project_points = X_pca
        if len(project_points) < 10:
            continue
        pcd = o3d.geometry.PointCloud()
        pcd.points = o3d.utility.Vector3dVector(project_points)
        pcd.colors = pcd.colors

        h,w=leaf_high_width(pcd)
        print(h,w)
        if h > 0.5 or w > 0.5:
            continue
        h1.append(h*10)
        w1.append(w*10)
    print(h1)
    hd = detect_outliers2(list(h1))
    wd = detect_outliers2(list(w1))
    print(itemy,np.max(hd),np.max(wd),idx)
    hs=sum(hd)
    ws=sum(wd)
    lenth.append(hd)
    width.append(wd)




import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
data1=lenth
id=1
for item in range(len(data1)):
    index = id + item
    for i in range(len(data1[item])):
        sheet.cell(row=i + 1, column=index, value=data1[item][i])


wb.save('End2End叶长.xlsx')

wb1 = openpyxl.Workbook()
sheet = wb1.active
data=width
id=1
for item in range(len(data)):
    index = id + item
    for i in range(len(data[item])):
        sheet.cell(row=i + 1, column=index, value=data[item][i])


wb1.save('End2End叶宽.xlsx')



